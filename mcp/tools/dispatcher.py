"""
mcp_gateway/mcp/tools/dispatcher.py

Tool call dispatcher — routes calls to Odoo ORM, HTTP APIs, or MCP servers.
Supports the full 18 specs representing Odoo ORM, file lookups, safe_eval scripting,
visual graphs, webapps, hot-loading modules, and subagents.
"""

import logging
import json
import re
import os
import requests
import base64
import urllib.request as _urllib_req
import urllib.parse as _urllib_parse
import socket
import ipaddress
import hashlib
import hmac
import math
import itertools
from html import escape as _html_escape
import datetime as dt_module
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter
import threading
from lxml import etree

from odoo.fields import Command
from odoo import exceptions
from odoo.tools.safe_eval import safe_eval, wrap_module

_logger = logging.getLogger(__name__)

_safe_json = wrap_module(json, ['dumps', 'loads', 'JSONDecodeError'])
_safe_re = wrap_module(re, ['compile', 'match', 'search', 'findall', 'finditer', 'sub', 'split',
                             'escape', 'fullmatch', 'IGNORECASE', 'I', 'MULTILINE', 'M',
                             'DOTALL', 'S', 'VERBOSE', 'X', 'error'])
_safe_math = wrap_module(math, ['floor', 'ceil', 'sqrt', 'log', 'log2', 'log10', 'exp', 'pow',
                                 'fabs', 'factorial', 'gcd', 'pi', 'e', 'inf', 'nan',
                                 'isnan', 'isinf', 'isfinite', 'trunc', 'degrees', 'radians',
                                 'sin', 'cos', 'tan', 'atan', 'atan2', 'hypot'])
_safe_itertools = wrap_module(itertools, [
    'chain', 'groupby', 'islice', 'product', 'combinations',
    'permutations', 'starmap', 'takewhile', 'dropwhile', 'zip_longest',
])
_safe_hashlib = wrap_module(hashlib, ['md5', 'sha1', 'sha256', 'sha512', 'new', 'pbkdf2_hmac'])
_safe_hmac = wrap_module(hmac, ['new', 'digest', 'compare_digest'])

import io as _io
import xlrd as _xlrd
import openpyxl as _openpyxl

_safe_base64 = wrap_module(base64, ['b64encode', 'b64decode', 'encodebytes', 'decodebytes'])
_safe_io = wrap_module(_io, ['BytesIO', 'StringIO'])
_safe_xlrd = wrap_module(_xlrd, ['open_workbook', 'xldate_as_datetime'])
_safe_openpyxl = wrap_module(_openpyxl, ['load_workbook'])

# Any avatar_* size field auto-redirects to image_1920 (the one field Odoo actually stores)
_AVATAR_FIELD_RE = re.compile(r'^avatar_(128|256|512|1024|1920)$')

# Old model names renamed across Odoo versions — remapped so a stale/guessed name still resolves
_MODEL_HISTORY = {
    'account.invoice': 'account.move',
    'account.invoice.line': 'account.move.line',
    'hr.holidays': 'hr.leave',
    'hr.holidays.status': 'hr.leave.type',
    'hr.holidays.allocation': 'hr.leave.allocation',
    'stock.quant.move': 'stock.move',
    'crm.case.categ': 'crm.tag',
    'product.attribute.value': 'product.attribute.value',
    'sale.order.line': 'sale.order.line',
}

_BULK_MAX = 1000  # flat cap for bulk create/update/delete operations


def _make_serializable(obj):
    """Recursively convert ORM results to JSON-safe types."""
    if isinstance(obj, (dt_module.date, dt_module.datetime)):
        return obj.isoformat()
    if isinstance(obj, tuple):  # Many2one from read_group returns (id, 'name')
        return [_make_serializable(x) for x in obj]
    if isinstance(obj, list):
        return [_make_serializable(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _make_serializable(v) for k, v in obj.items()}
    if hasattr(obj, 'ids'):  # recordset
        return obj.ids
    if isinstance(obj, bytes):  # binary fields (images, attachments)
        return None
    return obj


def _safe_exec(code: str, context: dict, fn_name: str = '_exec_fn'):
    """
    Execute multi-statement Python code inside safe_eval and return its result.

    Wraps `code` in a ``def <fn_name>():`` function so assignments and other
    statements work (safe_eval defaults to expression-only mode).  Uses AST to
    detect the last top-level expression and automatically inserts ``return``
    before it, so callers get the result without having to add ``return``
    themselves.

    Args:
        code (str): Python source to execute (may contain assignments, loops, etc.).
        context (dict): safe_eval globals/locals dict (modified in place by exec).
        fn_name (str): Name of the wrapper function injected into context.

    Returns:
        Any: The return value of the last expression in ``code``.
    """
    import ast as _ast
    stripped = code.strip()
    try:
        _tree = _ast.parse(stripped)
        if _tree.body and isinstance(_tree.body[-1], _ast.Expr):
            _lines = stripped.splitlines()
            _start = _tree.body[-1].lineno - 1  # 0-indexed
            _lines[_start] = 'return ' + _lines[_start]
            stripped = '\n'.join(_lines)
    except SyntaxError:
        pass  # safe_eval will surface the real error
    indented = '\n'.join('    ' + line for line in stripped.splitlines())
    fn_code = f'def {fn_name}():\n{indented}\n'
    safe_eval(fn_code, context, mode='exec', nocopy=True)
    return context[fn_name]()


def _render_via_odoo_report_layout(env, content_html, landscape=False):
    """Wrap ad-hoc HTML in Odoo's own configured report chrome (company logo,
    address, footer, page numbers — Settings > Companies > Document Layout)
    and return PDF bytes. Reuses `web.preview_externalreport`'s doc-less
    pattern (the same template behind the "Preview External Report" button)
    — no ir.actions.report record or QWeb view file needed for this: company/
    res_company are auto-defaulted by _render_template, and _prepare_html
    works fine on a bare `ir.actions.report` recordset with no real report."""
    Report = env['ir.actions.report']
    tmpl = etree.fromstring(f'''<t t-call="web.html_container">
        <t t-call="web.external_layout">
            <div class="page">{content_html}</div>
        </t>
    </t>''')
    full_html = Report._render_template(tmpl, {'report_type': 'pdf'})
    bodies, _res_ids, header, footer, paperformat_args = Report._prepare_html(full_html)
    return Report._run_wkhtmltopdf(
        bodies, header=header, footer=footer,
        specific_paperformat_args=paperformat_args, landscape=landscape,
    )


def _assert_public_url(url):
    """Reject URLs that resolve to internal/loopback/link-local addresses (SSRF guard).

    ponytail: resolve-then-connect has a small DNS-rebinding TOCTOU window (the
    hostname could re-resolve to a different IP between this check and urlopen).
    Pinning the checked IP for the actual connection would close it; add if this
    tool is ever exposed to less-trusted callers than an already-authenticated agent.
    """
    parsed = _urllib_parse.urlparse(url)
    if parsed.scheme not in ('http', 'https'):
        raise ValueError(f'URL scheme must be http or https, got: {parsed.scheme!r}')
    if not parsed.hostname:
        raise ValueError('URL has no hostname')
    try:
        addrs = socket.getaddrinfo(parsed.hostname, None)
    except socket.gaierror as e:
        raise ValueError(f'Could not resolve host {parsed.hostname!r}: {e}')
    for family, _type, _proto, _canon, sockaddr in addrs:
        ip = ipaddress.ip_address(sockaddr[0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast or ip.is_unspecified:
            raise ValueError(f'URL host {parsed.hostname!r} resolves to a non-public address ({ip}) — refusing to fetch')


_BLOCKED_RECORDSET_METHODS = frozenset({
    'write', 'unlink', 'create', 'copy', 'sudo', 'su', 'with_user', 'with_context',
    'toggle_active', 'action_archive', 'action_unarchive', 'message_post', 'message_post_with_source',
})
_BLOCKED_ENV_ATTRS = frozenset({'cr', 'registry', 'sudo', 'su'})


def _wrap_readonly(value):
    """If `value` is an Odoo recordset (has the shape of one), wrap it so any
    mutating call raises instead of executing for real. Plain values (str,
    int, list of dicts from read_group, etc.) pass through untouched."""
    if hasattr(value, '_name') and hasattr(value, 'browse'):
        return _ReadOnlyRecordset(value)
    return value


class _ReadOnlyRecordset:
    """Read-only view of an Odoo recordset for AI-authored `data_code` in
    create_echart/generate_export_file. safe_eval's opcode blacklist stops
    imports and attribute-assignment but NOT ordinary method calls, so
    `env['x'].search([]).unlink()` would otherwise execute for real from
    inside a tool that's only supposed to compute chart/export data.

    ponytail: no __add__/__eq__/__and__ — recordset concatenation/equality
    isn't used by existing chart/export data_code; add if a real case needs it.
    """
    __slots__ = ('_rs',)

    def __init__(self, rs):
        self._rs = rs

    def __getattr__(self, name):
        if name in _BLOCKED_RECORDSET_METHODS:
            raise AttributeError(
                f"'{name}' is blocked here — this tool can only read data, not write it. "
                f"Use execute_orm if a real write is actually needed."
            )
        value = getattr(self._rs, name)
        if not callable(value):
            return _wrap_readonly(value)
        def _wrapped(*args, **kwargs):
            return _wrap_readonly(value(*args, **kwargs))
        return _wrapped

    def __iter__(self):
        return (_wrap_readonly(r) for r in self._rs)

    def __len__(self):
        return len(self._rs)

    def __bool__(self):
        return bool(self._rs)

    def __getitem__(self, key):
        return _wrap_readonly(self._rs[key])

    def __repr__(self):
        return repr(self._rs)


class _ReadOnlyEnv:
    """Read-only view of an Odoo Environment — see _ReadOnlyRecordset."""
    __slots__ = ('_env',)

    def __init__(self, env):
        self._env = env

    def __getitem__(self, model_name):
        return _ReadOnlyRecordset(self._env[model_name])

    def __getattr__(self, name):
        if name in _BLOCKED_ENV_ATTRS:
            raise AttributeError(
                f"'{name}' is blocked here — direct cursor/registry/sudo access "
                f"would bypass the read-only guard entirely."
            )
        value = getattr(self._env, name)
        if not callable(value):
            return _wrap_readonly(value)
        def _wrapped(*args, **kwargs):
            return _wrap_readonly(value(*args, **kwargs))
        return _wrapped


def _normalize_domain(domain):
    # LLMs sometimes double-wrap: [[["f","=","v"]]] → [["f","=","v"]]
    if domain and isinstance(domain[0], list) and domain[0] and isinstance(domain[0][0], list):
        return domain[0]
    return domain


def _check_domain_fields(model, domain):
    # Odoo silently drops (treats as always-true) domain leaves on non-stored
    # computed fields without a custom search — LLMs then either see unfiltered
    # results or waste turns retrying the same broken domain. Fail loud instead.
    for leaf in domain:
        if not isinstance(leaf, (list, tuple)) or len(leaf) != 3:
            continue  # skip '&'/'|'/'!' operators
        fname = leaf[0]
        if not isinstance(fname, str) or '.' in fname:
            continue  # skip related/dotted paths
        field = model._fields.get(fname)
        if field is not None and not field.store and not field.search:
            raise ValueError(
                f"Field '{fname}' on {model._name} is a non-stored computed field and cannot be "
                f"used in a search domain. Use read_group on the model that stores this data instead "
                f"(e.g. group sale.order by partner_id rather than filtering res.partner.sale_order_count)."
            )


def _check_order_fields(model, order):
    # Odoo search/read fails with a database-level SQL conversion traceback if
    # the order clause references a non-stored computed field. Fail loud instead.
    if not order or not isinstance(order, str):
        return
    for part in order.split(','):
        part = part.strip()
        if not part:
            continue
        fname = part.split()[0]
        if '.' in fname:
            continue  # skip related/dotted paths
        field = model._fields.get(fname)
        if field is not None and not field.store:
            raise ValueError(
                f"Field '{fname}' on {model._name} is a non-stored computed field and cannot be "
                f"used in an order clause. Sort the results in Python after reading or sort by a stored field instead."
            )



class ToolDispatcher:
    """
    Routes tool calls to the correct backend execution environment.
    """

    def dispatch(self, tool, arguments: dict, env, user) -> str:
        """
        Execute a tool call and return result as JSON string.
        """
        _logger.info('DISPATCHER.dispatch called: tool=%s, args=%s', tool.name, arguments)
        try:
            if tool.tool_type == 'odoo':
                return self._dispatch_odoo(tool, arguments, env, user)
            elif tool.tool_type == 'external':
                return self._dispatch_http(tool, arguments)
            elif tool.tool_type == 'mcp_server':
                return self._dispatch_mcp_server(tool, arguments)
            else:
                raise ValueError(f'Unknown tool type: {tool.tool_type}')

        except Exception as e:
            _logger.error('Tool dispatch failed for %s: %s', tool.name, str(e))
            return json.dumps({
                'success': False,
                'error': str(e)[:500],
            })

    def _dispatch_odoo(self, tool, arguments: dict, env, user) -> str:
        """
        Execute tool on Odoo ORM.
        """
        if isinstance(arguments, str):
            try:
                arguments = json.loads(arguments)
            except (json.JSONDecodeError, TypeError) as e:
                return json.dumps({'success': False, 'error': f'Invalid arguments format: {str(e)}'})

        if not isinstance(arguments, dict):
            return json.dumps({'success': False, 'error': f'Arguments must be a dict, got: {type(arguments).__name__}'})

        try:
            # ── Route by tool name directly (18 generic tools) ───────────────────
            name = tool.name
            result = None

            if name == 'list_models':
                result = self._dispatch_list_models(arguments, env, user)
            elif name == 'get_model_schema':
                result = self._dispatch_get_model_schema(arguments, env, user)
            elif name == 'search_read':
                result = self._dispatch_search_read(arguments, env, user)
            elif name == 'read_record':
                result = self._dispatch_read_record(arguments, env, user)
            elif name == 'read_group':
                result = self._dispatch_read_group(arguments, env, user)
            elif name == 'create_record':
                result = self._dispatch_create_record(arguments, env, user)
            elif name == 'update_record':
                result = self._dispatch_update_record(arguments, env, user)
            elif name == 'delete_record':
                result = self._dispatch_delete_record(arguments, env, user)
            elif name == 'execute_method':
                result = self._dispatch_execute_method(arguments, env, user)
            elif name == 'execute_orm':
                result = self._dispatch_execute_orm(arguments, env, user)
            elif name == 'post_message':
                result = self._dispatch_post_message(arguments, env, user)
            elif name == 'get_attachments':
                result = self._dispatch_get_attachments(arguments, env, user)
            elif name == 'read_attachment':
                result = self._dispatch_read_attachment(arguments, env, user)
            elif name == 'set_binary_field':
                result = self._dispatch_set_binary_field(arguments, env, user)
            elif name == 'upload_attachment':
                result = self._dispatch_upload_attachment(arguments, env, user)
            elif name == 'create_records':
                result = self._dispatch_create_records(arguments, env, user)
            elif name == 'update_records':
                result = self._dispatch_update_records(arguments, env, user)
            elif name == 'delete_records':
                result = self._dispatch_delete_records(arguments, env, user)
            elif name == 'lookup_model_history':
                result = self._dispatch_lookup_model_history(arguments, env, user)
            elif name == 'accounting_health_summary':
                result = self._dispatch_accounting_health_summary(arguments, env, user)
            elif name == 'import_from_file':
                result = self._dispatch_import_from_file(arguments, env, user)
            elif name == 'create_echart':
                result = self._dispatch_create_echart(arguments, env, user)
            elif name == 'generate_export_file':
                result = self._dispatch_generate_export_file(arguments, env, user)
            elif name == 'ai_agent_query':
                result = self._dispatch_ai_agent_query(arguments, env, user)
            else:
                # Fallback for old custom model mapping methods
                model_name = tool.odoo_model
                method_name = tool.odoo_method
                model = env[model_name].sudo() if tool.sudo_execute else env[model_name].with_user(user)
                method = getattr(model, method_name)
                result = method(**arguments)

            return json.dumps({'success': True, 'result': result})

        except Exception as e:
            _logger.error('ORM tool run failed: %s', str(e))
            return json.dumps({'success': False, 'error': str(e)})

    # ════════════════════════════════════════════════════════════════
    # INDIVIDUAL ORM DISPATCH METHODS
    # ════════════════════════════════════════════════════════════════

    def _dispatch_list_models(self, arguments, env, user):
        regex = arguments.get('regex')
        models = env['ir.model'].search([])
        if regex:
            pattern = re.compile(regex)
            models = models.filtered(lambda m: pattern.search(m.model))
        result = []
        for m in models:
            model_class = env.registry.get(m.model)
            if model_class and not model_class._abstract and not model_class._transient:
                result.append({'model': m.model, 'name': m.name})
        return result

    # Mixin field prefixes that are usually noise when the AI is looking for data fields
    _SCHEMA_NOISE_PREFIXES = ('activity_', 'message_', 'website_message_', 'mail_')

    def _dispatch_get_model_schema(self, arguments, env, user):
        model_name = arguments['model']
        fields = arguments.get('fields')
        model = env.get(model_name)
        if model is None:
            raise ValueError(f"Model {model_name} does not exist.")
        result = model.with_user(user).fields_get(fields, ['type', 'string', 'help', 'relation', 'selection'])
        if fields:
            return result
        # Put noisy mixin fields at the end so the AI finds data fields first
        primary = {k: v for k, v in result.items() if not any(k.startswith(p) for p in self._SCHEMA_NOISE_PREFIXES)}
        noise = {k: v for k, v in result.items() if any(k.startswith(p) for p in self._SCHEMA_NOISE_PREFIXES)}
        return {**primary, **noise}

    _AUDIT_FIELDS = frozenset(['create_uid', 'write_uid', 'create_date', 'write_date', '__last_update'])

    def _dispatch_search_read(self, arguments, env, user):
        model_name = arguments['model']
        domain = _normalize_domain(arguments.get('domain', []))
        fields = arguments.get('fields', [])
        limit = min(arguments.get('limit', 10), 100)
        order = arguments.get('order')
        explicit_fields = bool(fields)

        model = env[model_name].with_user(user)
        _check_domain_fields(model, domain)
        _check_order_fields(model, order)
        records = model.search_read(domain, fields, limit=limit, order=order)
        # Detect binary fields so we can swap bytes→URL instead of bytes→None
        binary_fields = {f for f in fields if f in model._fields and model._fields[f].type == 'binary'} if fields else set()
        result = []
        for rec in records:
            s = _make_serializable(rec)
            rec_id = s.get('id')
            if rec_id and binary_fields:
                for bf in binary_fields:
                    if s.get(bf) is None:  # bytes were stripped by _make_serializable
                        s[bf] = f'/web/image/{model_name}/{rec_id}/{bf}'
            # ponytail: strip audit noise when AI didn't specify fields — keeps responses clean
            if not explicit_fields:
                for af in self._AUDIT_FIELDS:
                    s.pop(af, None)
            result.append(s)
        return result

    def _dispatch_read_record(self, arguments, env, user):
        model_name = arguments['model']
        res_id = arguments['res_id']
        fields = arguments.get('fields', [])

        record = env[model_name].with_user(user).browse(res_id)
        if not record.exists():
            raise ValueError(f"Record {res_id} not found in model {model_name}.")

        return _make_serializable(record.read(fields)[0])

    def _dispatch_read_group(self, arguments, env, user):
        model_name = arguments['model']
        domain = _normalize_domain(arguments.get('domain', []))
        fields = arguments['fields']
        groupby = arguments['groupby']
        limit = arguments.get('limit')
        orderby = arguments.get('order') or arguments.get('orderby')

        model = env[model_name].with_user(user)
        _check_domain_fields(model, domain)
        kwargs = {'limit': limit}
        if orderby:
            kwargs['orderby'] = orderby
        rows = model.read_group(domain, fields, groupby, **kwargs)
        return [_make_serializable(row) for row in rows]

    def _dispatch_create_record(self, arguments, env, user):
        model_name = arguments['model']
        values = arguments['values']
        match_field = arguments.get('match_field')

        if match_field and match_field in values:
            existing = env[model_name].with_user(user).search(
                [(match_field, '=', values[match_field])], limit=1
            )
            if existing:
                return {'id': existing.id, 'model': model_name, 'skipped': True,
                        'reason': f'{match_field}={values[match_field]!r} already exists'}

        processed_vals = self._prepare_create_values(values, model_name, env)
        if isinstance(processed_vals, dict) and 'error' in processed_vals:
            raise ValueError(processed_vals['error'])

        model = env[model_name].with_user(user)
        record = model.create(processed_vals)
        return {'id': record.id, 'model': model_name}

    def _dispatch_update_record(self, arguments, env, user):
        model_name = arguments['model']
        res_ids = arguments['res_ids']
        values = arguments['values']
        
        processed_vals = self._prepare_create_values(values, model_name, env)
        if isinstance(processed_vals, dict) and 'error' in processed_vals:
            raise ValueError(processed_vals['error'])
            
        records = env[model_name].with_user(user).browse(res_ids)
        records.write(processed_vals)
        return True

    def _dispatch_delete_record(self, arguments, env, user):
        model_name = arguments['model']
        res_ids = arguments['res_ids']
        
        records = env[model_name].with_user(user).browse(res_ids)
        return records.unlink()

    def _dispatch_execute_method(self, arguments, env, user):
        model_name = arguments['model']
        res_ids = arguments['res_ids']
        method_name = arguments['method']
        args = arguments.get('args', [])
        kwargs = arguments.get('kwargs', {})
        
        if method_name.startswith('_'):
            raise PermissionError("Private method execution is strictly forbidden.")
            
        records = env[model_name].with_user(user).browse(res_ids)
        if not hasattr(records, method_name):
            raise AttributeError(f"Method {method_name} does not exist on {model_name}.")

        result = getattr(records, method_name)(*args, **kwargs)
        return _make_serializable(result)

    def _dispatch_post_message(self, arguments, env, user):
        model = arguments['model']
        record_id = int(arguments['record_id'])
        body = arguments['body']
        message_type = arguments.get('message_type', 'comment')
        subtype_xmlid = 'mail.mt_note' if message_type == 'internal' else 'mail.mt_comment'
        record = env[model].browse(record_id)
        msg = record.message_post(body=body, message_type=message_type, subtype_xmlid=subtype_xmlid)
        return {'message_id': msg.id, 'model': model, 'record_id': record_id}

    def _dispatch_get_attachments(self, arguments, env, user):
        model = arguments['model']
        record_id = int(arguments['record_id'])
        attachments = env['ir.attachment'].search_read(
            [('res_model', '=', model), ('res_id', '=', record_id)],
            ['id', 'name', 'mimetype', 'file_size', 'create_date']
        )
        for att in attachments:
            att['url'] = f'/web/content/{att["id"]}?download=true'
        # create_date comes back as a raw datetime from search_read — must
        # serialize before this hits the outer json.dumps().
        return _make_serializable({'model': model, 'record_id': record_id, 'attachments': attachments, 'count': len(attachments)})

    def _dispatch_upload_attachment(self, arguments, env, user):
        att = env['ir.attachment'].create({
            'name': arguments['filename'],
            'res_model': arguments['model'],
            'res_id': int(arguments['record_id']),
            'mimetype': arguments.get('mimetype', 'application/octet-stream'),
            'datas': arguments['datas'],
        })
        return {'id': att.id, 'name': att.name, 'model': arguments['model'], 'record_id': int(arguments['record_id'])}

    _MAX_ATTACHMENT_BYTES = 1 * 1024 * 1024  # 1 MiB cap — keeps large binary data out of the LLM context

    def _dispatch_read_attachment(self, arguments, env, user):
        att_id = int(arguments['attachment_id'])
        include_data = arguments.get('include_data', True)
        rows = env['ir.attachment'].search_read(
            [('id', '=', att_id)],
            ['id', 'name', 'mimetype', 'file_size', 'type', 'url', 'res_model', 'res_id']
        )
        if not rows:
            raise ValueError(f'Attachment not found: {att_id}')
        att = rows[0]
        data_base64 = None
        warnings = []
        file_size = int(att.get('file_size') or 0)
        if include_data and att.get('type') != 'url':
            if file_size > self._MAX_ATTACHMENT_BYTES:
                warnings.append(f'File is {file_size} bytes, over 1MB cap. Use the URL to download.')
            else:
                datas_rows = env['ir.attachment'].search_read([('id', '=', att_id)], ['datas'])
                raw = datas_rows[0].get('datas') if datas_rows else None
                if raw:
                    # ir.attachment.datas comes back as base64-encoded bytes, not str —
                    # json.dumps can't serialize bytes, decode (safe, base64 is ASCII).
                    data_base64 = raw.decode('ascii') if isinstance(raw, bytes) else raw
        elif att.get('type') == 'url':
            warnings.append('URL-type attachment — use the url field directly.')
        att['url'] = f'/web/content/{att_id}?download=true'
        return {
            'attachment': att,
            'data_base64': data_base64,
            'data_included': data_base64 is not None,
            'warnings': warnings,
        }

    def _dispatch_set_binary_field(self, arguments, env, user):
        """Fetch bytes from a URL and write them to a binary/image field. The bytes
        never pass through the LLM — only the source URL and target field do."""
        model = arguments['model']
        record_id = int(arguments['record_id'])
        field_name = arguments['field_name']
        source_url = arguments['source']

        if _AVATAR_FIELD_RE.match(field_name):
            field_name = 'image_1920'

        field_meta = env[model].fields_get([field_name], ['type', 'readonly'])
        if field_name not in field_meta:
            raise ValueError(f'Field {field_name} not found on {model}')
        if field_meta[field_name].get('type') not in ('binary', 'image'):
            raise ValueError(f'Field {field_name} is not a binary/image field')
        if field_meta[field_name].get('readonly'):
            raise ValueError(f'Field {field_name} is readonly')

        _assert_public_url(source_url)

        # 25MB cap, streamed in 64KB chunks
        _MAX_BINARY = 25 * 1024 * 1024
        chunks = []
        total = 0
        req = _urllib_req.Request(source_url, headers={'User-Agent': 'mcp-gateway/1.0'})
        with _urllib_req.urlopen(req, timeout=30) as resp:
            while True:
                chunk = resp.read(65536)
                if not chunk:
                    break
                total += len(chunk)
                if total > _MAX_BINARY:
                    raise ValueError('Source file exceeds 25MB limit')
                chunks.append(chunk)

        datas = base64.b64encode(b''.join(chunks)).decode('ascii')
        env[model].browse(record_id).write({field_name: datas})
        return {'model': model, 'record_id': record_id, 'field': field_name, 'size_bytes': total}

    def _dispatch_create_records(self, arguments, env, user):
        """Bulk create — multiple records in one call."""
        model_name = arguments['model']
        vals_list = arguments['vals_list']
        match_field = arguments.get('match_field')
        if len(vals_list) > _BULK_MAX:
            raise ValueError(f'vals_list exceeds limit of {_BULK_MAX}')

        def _coerce(vals):
            processed = self._prepare_create_values(vals, model_name, env)
            if isinstance(processed, dict) and 'error' in processed:
                raise ValueError(processed['error'])
            return processed

        model = env[model_name].with_user(user)
        if match_field:
            # Match on the raw (pre-coercion) values, same as the singular
            # create_record path — match_field is a plain identifier, not
            # something that needs date/many2one/m2m coercion for a search.
            match_values = [v[match_field] for v in vals_list if match_field in v]
            existing = {
                r[match_field]: r['id']
                for r in model.search_read([(match_field, 'in', match_values)], [match_field])
            } if match_values else {}
            to_create, skipped_ids = [], []
            for vals in vals_list:
                key = vals.get(match_field)
                if key and key in existing:
                    skipped_ids.append(existing[key])
                else:
                    to_create.append(_coerce(vals))
            new_ids = model.create(to_create).ids if to_create else []
            return {'model': model_name, 'ids': new_ids, 'skipped_ids': skipped_ids,
                    'count': len(new_ids), 'skipped_count': len(skipped_ids)}

        records = model.create([_coerce(vals) for vals in vals_list])
        return {'model': model_name, 'ids': records.ids, 'count': len(records)}

    def _dispatch_update_records(self, arguments, env, user):
        """Bulk update — the same values written to multiple records."""
        model_name = arguments['model']
        record_ids = arguments['record_ids']
        values = arguments['values']
        if len(record_ids) > _BULK_MAX:
            raise ValueError(f'record_ids exceeds limit of {_BULK_MAX}')
        processed_vals = self._prepare_create_values(values, model_name, env)
        if isinstance(processed_vals, dict) and 'error' in processed_vals:
            raise ValueError(processed_vals['error'])
        env[model_name].with_user(user).browse(record_ids).write(processed_vals)
        return {'model': model_name, 'record_ids': record_ids, 'count': len(record_ids)}

    def _dispatch_delete_records(self, arguments, env, user):
        """Bulk delete — multiple records unlinked in one call."""
        model = arguments['model']
        record_ids = arguments['record_ids']
        if len(record_ids) > _BULK_MAX:
            raise ValueError(f'record_ids exceeds limit of {_BULK_MAX}')
        env[model].with_user(user).browse(record_ids).unlink()
        return {'model': model, 'count': len(record_ids)}

    def _dispatch_lookup_model_history(self, arguments, env, user):
        """Resolve outdated/renamed model names to their current Odoo equivalent."""
        model = arguments['model']
        if model in _MODEL_HISTORY:
            current = _MODEL_HISTORY[model]
            return {
                'queried': model,
                'current_name': current,
                'renamed': current != model,
                'note': f'Use "{current}" instead of "{model}"' if current != model else f'"{model}" is current',
            }
        exists = bool(env.registry.get(model))
        return {
            'queried': model,
            'current_name': model if exists else None,
            'renamed': False,
            'exists': exists,
            'note': 'Model exists and name is current' if exists else f'Model "{model}" not found — check spelling or use list_models.',
        }

    def _dispatch_accounting_health_summary(self, arguments, env, user):
        """Quick accounts-receivable/payable health check without multiple tool calls."""
        Move = env['account.move'].with_user(user)
        today = date.today()  # compared against raw `date` values from search_read below — must stay a date, not a string
        ar = Move.search_read(
            [('move_type', '=', 'out_invoice'), ('state', '=', 'posted'),
             ('payment_state', 'not in', ['paid', 'reversed'])],
            ['id', 'name', 'partner_id', 'amount_residual', 'invoice_date_due']
        )
        ap = Move.search_read(
            [('move_type', '=', 'in_invoice'), ('state', '=', 'posted'),
             ('payment_state', 'not in', ['paid', 'reversed'])],
            ['id', 'name', 'partner_id', 'amount_residual', 'invoice_date_due']
        )
        draft = Move.search_count([('move_type', 'in', ('out_invoice', 'in_invoice')), ('state', '=', 'draft')])
        ar_overdue = [r for r in ar if r.get('invoice_date_due') and r['invoice_date_due'] < today]
        ap_overdue = [r for r in ap if r.get('invoice_date_due') and r['invoice_date_due'] < today]
        return {
            'receivables': {
                'total_count': len(ar), 'overdue_count': len(ar_overdue),
                'total_amount': sum(r['amount_residual'] for r in ar),
                'overdue_amount': sum(r['amount_residual'] for r in ar_overdue),
            },
            'payables': {
                'total_count': len(ap), 'overdue_count': len(ap_overdue),
                'total_amount': sum(r['amount_residual'] for r in ap),
                'overdue_amount': sum(r['amount_residual'] for r in ap_overdue),
            },
            'draft_invoice_backlog': draft,
            'as_of': today.isoformat(),
        }

    def _dispatch_import_from_file(self, arguments, env, user):
        """Parse a staged ir.attachment (CSV or Excel) and load rows into Odoo via env[model].load()."""
        import base64 as _b64
        import csv as _csv
        import io as _io

        attachment_id = int(arguments['attachment_id'])
        model = arguments['model']
        has_header = arguments.get('has_header', True)
        match_field = arguments.get('match_field')

        rows = env['ir.attachment'].search_read(
            [('id', '=', attachment_id)],
            ['name', 'mimetype', 'datas', 'file_size']
        )
        if not rows:
            raise ValueError(f'Attachment not found: {attachment_id}')
        att = rows[0]
        if not att.get('datas'):
            raise ValueError('Attachment has no data')

        raw = _b64.b64decode(att['datas'])
        mimetype = (att.get('mimetype') or '').lower()
        name = att.get('name', '')

        if (mimetype in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         'application/vnd.ms-excel')
                or name.lower().endswith(('.xlsx', '.xls'))):
            import openpyxl
            wb = openpyxl.load_workbook(filename=_io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            all_rows = [
                [str(cell.value) if cell.value is not None else '' for cell in row]
                for row in ws.iter_rows()
            ]
            wb.close()
        else:
            text = raw.decode('utf-8-sig')  # utf-8-sig strips BOM
            all_rows = list(_csv.reader(_io.StringIO(text)))

        if not all_rows:
            raise ValueError('File is empty')

        if has_header:
            fields = all_rows[0]
            data = all_rows[1:]
        else:
            fields = [f'col{i}' for i in range(len(all_rows[0]))]
            data = all_rows

        if not data:
            raise ValueError('File has no data rows (only a header row was found)')

        skipped_count = 0
        if match_field and match_field in fields:
            col_idx = fields.index(match_field)
            match_values = [row[col_idx] for row in data if len(row) > col_idx and row[col_idx]]
            existing_vals = {
                r[match_field]
                for r in env[model].with_user(user).search_read(
                    [(match_field, 'in', match_values)], [match_field]
                )
            } if match_values else set()
            filtered = [row for row in data if not (len(row) > col_idx and row[col_idx] in existing_vals)]
            skipped_count = len(data) - len(filtered)
            data = filtered

        result = env[model].with_user(user).load(fields, data)
        ids = result.get('ids') or []
        messages = result.get('messages') or []
        errors = [m for m in messages if m.get('type') in ('error', 'warning')]

        return {
            'model': model,
            'source_file': name,
            'fields': fields,
            'total_rows': len(data) + skipped_count,
            'created_count': len(ids),
            'skipped_count': skipped_count,
            'ids': ids[:100],
            'errors': errors[:20],
        }

    def _dispatch_execute_orm(self, arguments, env, user):
        code = arguments['code']

        # Strip ALL import lines — every needed module is pre-loaded in eval_context
        code = '\n'.join(
            line for line in code.splitlines()
            if not line.strip().startswith(('import ', 'from '))
        )

        def _read_excel(attachment_id, sheet=0):
            """Read xls/xlsx/ods attachment → list of row lists.
            attachment.datas is base64 in Odoo ORM; attachment.raw is the raw bytes."""
            import io as _real_io
            import zipfile as _zipfile
            import xml.etree.ElementTree as _ET
            att = env['ir.attachment'].browse(attachment_id)
            raw_b64 = att.datas
            if not raw_b64:
                raise ValueError(f"Attachment {attachment_id} has no data")
            raw = base64.b64decode(raw_b64)  # datas is ALWAYS base64 in Odoo ORM
            if raw[:4] == b'\xD0\xCF\x11\xE0':  # OLE2 magic → .xls
                wb = _xlrd.open_workbook(file_contents=raw)
                ws = wb.sheet_by_index(sheet)
                return [ws.row_values(r) for r in range(ws.nrows)]
            # ZIP-based: peek at content to distinguish xlsx vs ods
            try:
                with _zipfile.ZipFile(_real_io.BytesIO(raw)) as zf:
                    names = zf.namelist()
                if 'content.xml' in names:  # ODS
                    _NS_TABLE = 'urn:oasis:names:tc:opendocument:xmlns:table:1.0'
                    _NS_TEXT  = 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
                    with _zipfile.ZipFile(_real_io.BytesIO(raw)) as zf:
                        tree = _ET.parse(zf.open('content.xml'))
                    sheets = list(tree.getroot().iter(f'{{{_NS_TABLE}}}table'))
                    if not sheets:
                        return []
                    ws = sheets[min(sheet, len(sheets) - 1)]
                    rows = []
                    for row_el in ws.iter(f'{{{_NS_TABLE}}}table-row'):
                        cells = []
                        for cell in row_el:
                            tag = cell.tag.split('}')[-1]
                            if tag in ('table-cell', 'covered-table-cell'):
                                repeat = int(cell.get(f'{{{_NS_TABLE}}}number-columns-repeated', 1))
                                text = next((p.text or '' for p in cell.iter(f'{{{_NS_TEXT}}}p')), '')
                                cells.extend([text] * repeat)
                        # trim trailing blanks and skip empty rows
                        while cells and cells[-1] == '':
                            cells.pop()
                        if cells:
                            rows.append(cells)
                    return rows
                else:  # xlsx
                    wb = _openpyxl.load_workbook(_real_io.BytesIO(raw))
                    ws = wb.worksheets[min(sheet, len(wb.worksheets) - 1)]
                    return [list(row) for row in ws.iter_rows(values_only=True)]
            except _zipfile.BadZipFile:
                raise ValueError("Attachment is not a valid Excel/ODS file (bad ZIP)")

        eval_context = {
            'env': env,
            'user': user,
            'uid': user.id,
            'Command': Command,
            'json': _safe_json,
            're': _safe_re,
            'math': _safe_math,
            'itertools': _safe_itertools,
            'hashlib': _safe_hashlib,
            'hmac': _safe_hmac,
            'base64': _safe_base64,
            'io': _safe_io,
            'BytesIO': _safe_io.BytesIO,
            'StringIO': _safe_io.StringIO,
            'xlrd': _safe_xlrd,
            'openpyxl': _safe_openpyxl,
            'datetime': datetime,
            'date': date,
            'timedelta': timedelta,
            'defaultdict': defaultdict,
            'Counter': Counter,
            'read_excel': _read_excel,
            'print': print,
        }
        # Wrap in a function so multi-line code with assignments works.
        # _safe_exec handles the AST return-insertion and mode='exec' boilerplate.
        return _make_serializable(_safe_exec(code, eval_context, '_orm_fn'))

    def _dispatch_create_echart(self, arguments, env, user):
        name = arguments['name']
        data_code = arguments.get('data_code', '')
        options = arguments.get('options', {})

        if data_code:
            # env/user are wrapped read-only — data_code only ever needs to return
            # rows for the chart, never mutate records (see _ReadOnlyRecordset).
            fn_globals = {
                'env': _ReadOnlyEnv(env),
                'user': _wrap_readonly(user),
                'datetime': datetime,
                'date': date,
                'timedelta': timedelta,
                'defaultdict': defaultdict,
                'Counter': Counter,
                'json': _safe_json,
                're': _safe_re,
                'math': _safe_math,
            }
            # Strip ALL import lines — every needed module is pre-loaded in fn_globals above
            stripped_code = '\n'.join(
                line for line in data_code.strip().splitlines()
                if not line.strip().startswith(('import ', 'from '))
            )
            options = _safe_exec(stripped_code, fn_globals, '_chart_fn')

        chart = env['mcp.echart'].with_user(user).create({
            'name': name,
            'data_code': data_code,
            'options': json.dumps(options) if isinstance(options, dict) else (options or '{}'),
        })
        return {'id': chart.id, 'name': name}

    _MAX_EXPORT_ROWS = 10000  # ponytail: flat cap, raise if a real use case needs more

    def _dispatch_generate_export_file(self, arguments, env, user):
        """AI supplies data_code that returns rows (list of lists, header row first) —
        the sandbox never touches a file-writing library directly (safe_eval blocks
        STORE_ATTR entirely, so `ws.title = ...` always fails; and execute_orm's
        openpyxl binding is read-only). This method builds the actual file itself,
        outside the sandbox, from the AI-returned data only."""
        export_format = arguments['format']
        if export_format not in ('csv', 'xlsx', 'pdf'):
            raise ValueError(f"Unsupported format: {export_format!r} — use csv, xlsx, or pdf")
        title = arguments.get('title') or 'Export'
        data_code = arguments['data_code']

        # env/user are wrapped read-only — same reasoning as create_echart above.
        fn_globals = {
            'env': _ReadOnlyEnv(env),
            'user': _wrap_readonly(user),
            'datetime': datetime,
            'date': date,
            'timedelta': timedelta,
            'defaultdict': defaultdict,
            'Counter': Counter,
            'json': _safe_json,
            're': _safe_re,
            'math': _safe_math,
        }
        stripped_code = '\n'.join(
            line for line in data_code.strip().splitlines()
            if not line.strip().startswith(('import ', 'from '))
        )
        rows = _safe_exec(stripped_code, fn_globals, '_export_fn')

        if not isinstance(rows, list) or not rows or not all(isinstance(r, (list, tuple)) for r in rows):
            raise ValueError('data_code must return a non-empty list of rows (each row a list) — header row first')
        if len(rows) > self._MAX_EXPORT_ROWS:
            raise ValueError(f'{len(rows)} rows exceeds the {self._MAX_EXPORT_ROWS}-row export cap')

        filename = arguments.get('filename') or re.sub(r'[^A-Za-z0-9_-]+', '_', title).strip('_') or 'export'

        if export_format == 'csv':
            import csv
            buf = _io.StringIO()
            csv.writer(buf).writerows(rows)
            file_bytes = buf.getvalue().encode('utf-8')
            mimetype = 'text/csv'
            filename += '.csv'
        elif export_format == 'xlsx':
            wb = _openpyxl.Workbook()
            ws = wb.active
            for row in rows:
                ws.append(list(row))
            buf = _io.BytesIO()
            wb.save(buf)
            file_bytes = buf.getvalue()
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename += '.xlsx'
        else:  # pdf
            header_cells = ''.join(f'<th>{_html_escape(str(c))}</th>' for c in rows[0])
            body_rows = ''.join(
                '<tr>' + ''.join(f'<td>{_html_escape(str(c))}</td>' for c in row) + '</tr>'
                for row in rows[1:]
            )
            row_count = len(rows) - 1
            generated_at = datetime.now().strftime('%B %d, %Y at %H:%M')
            content_html = f'''
                <h2>{_html_escape(title)}</h2>
                <p class="text-muted">Generated {generated_at} · {row_count} row(s)</p>
                <table class="table table-sm table-bordered table-striped">
                    <thead><tr>{header_cells}</tr></thead>
                    <tbody>{body_rows}</tbody>
                </table>
            '''
            file_bytes = _render_via_odoo_report_layout(env, content_html)
            mimetype = 'application/pdf'
            filename += '.pdf'

        att = env['ir.attachment'].with_user(user).create({
            'name': filename,
            'datas': base64.b64encode(file_bytes).decode('ascii'),
            'mimetype': mimetype,
        })
        return {
            'id': att.id,
            'url': f'/web/content/{att.id}?download=true',
            'filename': filename,
            'mimetype': mimetype,
            'size_bytes': len(file_bytes),
        }

    def _dispatch_ai_agent_query(self, arguments, env, user):
        agent_id = arguments['agent_id']
        prompt = arguments['prompt']
        
        if not hasattr(threading.current_thread(), 'mcp_depth'):
            threading.current_thread().mcp_depth = 0
            
        if threading.current_thread().mcp_depth >= 3:
            raise exceptions.ValidationError("Maximum subagent query delegation depth exceeded (loop detected).")
            
        threading.current_thread().mcp_depth += 1
        try:
            agent = env['mcp.agent'].with_user(user).browse(agent_id)
            if not agent.exists():
                raise ValueError(f"Agent {agent_id} does not exist.")
            from ..gateway import McpGateway  # lazy import — avoids circular dependency
            gateway = McpGateway(env, user)
            result = gateway.run(agent_id=agent_id, user_message=prompt)
            return result.get('reply', '')
        finally:
            threading.current_thread().mcp_depth -= 1

    # ════════════════════════════════════════════════════════════════
    # LEGACY ORM VALUE FORMATTERS
    # ════════════════════════════════════════════════════════════════

    def _dispatch_http(self, tool, arguments: dict) -> str:
        try:
            url = tool.endpoint_url
            method = tool.http_method
            headers = self._build_auth_headers(tool)
            timeout = tool.timeout_seconds

            if method == 'GET':
                response = requests.get(url, params=arguments, headers=headers, timeout=timeout)
            elif method == 'POST':
                response = requests.post(url, json=arguments, headers=headers, timeout=timeout)
            elif method == 'PUT':
                response = requests.put(url, json=arguments, headers=headers, timeout=timeout)
            elif method == 'DELETE':
                response = requests.delete(url, params=arguments, headers=headers, timeout=timeout)
            else:
                raise ValueError(f'Unsupported HTTP method: {method}')

            response.raise_for_status()
            data = response.json()

            if tool.response_path:
                result = self._extract_path(data, tool.response_path)
            else:
                result = data

            return json.dumps({'success': True, 'result': result})

        except Exception as e:
            return json.dumps({'success': False, 'error': str(e)[:500]})

    def _dispatch_mcp_server(self, tool, arguments: dict) -> str:
        try:
            url = f'{tool.mcp_server_url}/call'
            headers = {
                'Authorization': f'Bearer {tool.mcp_server_key}',
                'Content-Type': 'application/json',
            }
            payload = {
                'tool': tool.name,
                'arguments': arguments,
            }

            response = requests.post(url, json=payload, headers=headers, timeout=30)
            response.raise_for_status()

            return json.dumps({'success': True, 'result': response.json()})

        except Exception as e:
            return json.dumps({'success': False, 'error': str(e)[:500]})

    def _extract_path(self, data: dict, path: str):
        parts = path.split('.')
        current = data
        for part in parts:
            if isinstance(current, dict):
                current = current.get(part)
            elif isinstance(current, list):
                try:
                    current = current[int(part)]
                except (ValueError, IndexError):
                    return None
            else:
                return None
        return current

    def _build_auth_headers(self, tool) -> dict:
        headers = {}
        if tool.auth_type == 'none':
            pass
        elif tool.auth_type == 'bearer':
            headers['Authorization'] = f'Bearer {tool.auth_value}'
        elif tool.auth_type == 'basic':
            encoded = base64.b64encode(tool.auth_value.encode()).decode()
            headers['Authorization'] = f'Basic {encoded}'
        elif tool.auth_type == 'api_key_header':
            headers[tool.auth_header_name] = tool.auth_value
        return headers

    def _prepare_create_values(self, arguments, model_name: str, env) -> dict:
        if isinstance(arguments, str):
            try:
                arguments = json.loads(arguments)
            except (json.JSONDecodeError, TypeError) as e:
                return {'error': f'Invalid arguments format: {str(e)}'}

        if not isinstance(arguments, dict):
            return {'error': f'Arguments must be a dict, got: {type(arguments).__name__}'}

        vals = {}
        model = env.get(model_name)
        if model is None:
            return {'error': f'Model {model_name} not found'}

        # ponytail: remap known Odoo version field renames before validation
        _FIELD_RENAMES = {'detailed_type': 'type'}
        arguments = {_FIELD_RENAMES.get(k, k): v for k, v in arguments.items()}

        for key, value in arguments.items():
            if value is None:
                continue

            field = model._fields.get(key)
            if not field:
                vals[key] = value
                continue

            if field.type == 'datetime' and isinstance(value, str):
                value = self._parse_datetime(value)
            elif field.type == 'date' and isinstance(value, str):
                value = self._parse_date(value)
            elif field.type == 'many2one':
                if isinstance(value, list):
                    value = value[0] if value else False
                elif isinstance(value, str) and value.isdigit():
                    value = int(value)
            elif field.type in ('one2many', 'many2many'):
                if isinstance(value, list):
                    # Handle raw Odoo Command tuples directly
                    # e.g., [[0, 0, {...}]] or [[6, 0, [1, 2]]]
                    if value and isinstance(value[0], list):
                        # Construct Command tuples
                        cmd_list = []
                        for item in value:
                            if len(item) == 3:
                                cmd_list.append((item[0], item[1], item[2]))
                        value = cmd_list
                    elif value and isinstance(value[0], int):
                        value = [Command.set(value)]
                elif isinstance(value, str) and value:
                    try:
                        ids = [int(x.strip()) for x in value.split(',') if x.strip().isdigit()]
                        value = [Command.set(ids)]
                    except ValueError:
                        pass
            vals[key] = value
        return vals

    def _parse_datetime(self, value: str):
        if not isinstance(value, str):
            return value
        value = re.sub(r'[+-]\d{2}:?\d{2}$', '', value)
        value = re.sub(r'Z$', '', value)
        value = value.strip()
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M',
            '%Y-%m-%d',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        return value

    def _parse_date(self, value: str):
        if not isinstance(value, str):
            return value
        formats = [
            '%Y-%m-%d',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return value
